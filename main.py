from openpyxl import load_workbook, Workbook
import requests


def update_excel_results(file_path=None):
    # 读取excel
    wb = load_workbook(r'C:\Users\liu51\Desktop\Test\00 资料\9_python自动化\testcase_api_wuye.xlsx')

    # 获取工作表
    sheet = wb['register']

    # 获取最大行数
    print(dir(sheet))
    max_row = sheet.max_row
    # 读取所有字段数据
    for row in sheet.iter_rows(min_row=2, max_row=max_row):
        case_id = row[0].value
        intelface = row[1].value
        title = row[2].value
        method = row[3].value
        url = row[4].value
        data = row[5].value
        expcted = row[6].value
        result_cell = row[7].value

        # 发送http请求并获取结果
        result = send_request(method, url, data)

        # 将结果写入result列
        result.value = result

    # 保存修改后表格
    wb.save(file_path)

def send_request(method, url, data):
    headers = {
        'Content-Type': 'application/json',
    }

    if method.upper() == 'POST':
        response = requests.post(url, headers=headers, json=eval(data))
    elif method.upper() == 'GET':
        response = requests.get(url, headers=headers, params=eval(data))
    else:
        return f"Unsupported method: {method}"

    if response.status_code == 200:
        return response.json()
    else:
        return f"Error {response.status_code}: {response.text}"



    # 测试

update_excel_results()
